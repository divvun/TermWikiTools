# -*- coding: utf-8 -*-
"""Check if expressions already exist in TermWiki."""
import argparse
import string
from pathlib import Path

from lxml import etree

from openpyxl import Workbook
from termwikiimporter import importer, lookup, read_termwiki


def languages(wikitree):
    return {
        related_expression['language']
        for concept in concepts(wikitree)
        for related_expression in concept.related_expressions
    }


def language_positions(languages):
    return {language: x for x, language in enumerate(languages, start=1)}


def concepts(wikitree):
    for wikipage in wikitree.xpath('.//page'):
        concept = read_termwiki.Concept()
        concept.title = wikipage.get('title')
        concept.from_termwiki(wikipage.find('./concept').text)
        yield concept


def concept_hits(concept):
    return {
        hit.replace(' ', '_')
        for related_expression in concept.related_expressions
        for hit in lookup.lookup(related_expression['expression'],
                                 related_expression['language'])
    }



def hitlist(wikitree, report_file):
    d = dict(enumerate(string.ascii_uppercase, 1))
    dunks = [(concept, concept_hits(concept))
             for concept in concepts(wikitree)]

    l = languages(wikitree)
    lf = language_positions(l)

    wb = Workbook()
    ws = wb.active

    ws.cell(
        row=1,
        column=1,
        value=
        f'Possible duplicate concepts: {len([dunk[1] for dunk in dunks if dunk[1]])} of totally {len(dunks)}'
    )

    for language in l:
        ws.cell(row=2, column=lf[language], value=language)

    for row, dunk in enumerate(dunks, start=3):
        for language in l:
            v = ', '.join([
                related_expression['expression']
                for related_expression in dunk[0].related_expressions
                if related_expression['language'] == language
            ])
            ws.cell(row=row, column=lf[language], value=v)
            ws.column_dimensions[d[lf[language]]].width = 30
        for x, hit in enumerate(dunk[1], start=len(lf) + 1):
            ws.cell(row=2, column=x, value='Possible dupe')
            ws.cell(
                row=row,
                column=x,
                value=
                f'=HYPERLINK("https://satni.uit.no/termwiki/index.php?title={hit}", "{hit}"'
            )
            ws.column_dimensions[d[x]].width = 30

    wb.save(filename=report_file)
    print(f'Check {report_file}')


def parse_options():
    """Parse options given to the script."""
    parser = argparse.ArgumentParser(
        description='Look for possible dupes in Excel files.')

    parser.add_argument(
        'termfile',
        help='An Excel file containing terms. Each file must have a '
        'yaml file that inform how they should be treated.')

    args = parser.parse_args()

    return args


def main():
    args = parse_options()
    print(f'Checking for dupes in {args.termfile}')
    excel_importer = importer.init_file(args.termfile)
    p = Path(args.termfile)
    hitlist(excel_importer.get_concepts(), p.stem + '.report.xlsx')
