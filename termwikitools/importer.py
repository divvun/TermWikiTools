# -*- coding: utf-8 -*-
#
#   This program is free software: you can redistribute it and/or modify
#   it under the terms of the GNU General Public License as published by
#   the Free Software Foundation, either version 3 of the License, or
#   (at your option) any later version.
#
#   This program is distributed in the hope that it will be useful,
#   but WITHOUT ANY WARRANTY; without even the implied warranty of
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#   GNU General Public License for more details.
#
#   You should have received a copy of the GNU General Public License
#   along with this file. If not, see <http://www.gnu.org/licenses/>.
#
#   Copyright © 2016-2024 The University of Tromsø
#   http://giellatekno.uit.no & http://divvun.no
#
"""Convert term files to termwiki parsable xml."""

import json
from dataclasses import asdict
from pathlib import Path
from typing import Iterator

import click
import openpyxl
from marshmallow.exceptions import ValidationError

from termwikitools.read_termwiki import (
    COLLECTION_SCHEMA,
    TERMWIKI_PAGE_SCHEMA,
    cleanup_termwiki_page,
)


class SheetImporter:
    def __init__(self, sheet, comma: bool, lowercase: bool):
        self.sheet = sheet
        self.comma = comma
        self.lowercase = lowercase

    def row_to_concept_dict(
        self, collection: str, sheet_info: dict, row_number: int
    ) -> dict:
        return {
            "title": sheet_info.get("main_category"),
            "concept": {"collection": [collection]},
            "related_expressions": [
                related_expression
                for related_expression in self.make_expressions(
                    sheet_info.get("related_expressions") or [], row_number
                )
                if related_expression.get("expression")
            ],
            "concept_infos": (
                self.make_dict(sheet_info.get("concept_infos") or [], row_number)
                if sheet_info.get("concept_infos")
                else None
            ),
        }

    def make_expression_list(self, value: str | None) -> list[str]:
        if value is None:
            return []

        if self.lowercase:
            value = value.lower()

        return (
            [part.strip() for part in value.split(",")]
            if self.comma
            else [value.strip()]
        )

    def make_expressions(self, dict_templates: list, row_number: int) -> Iterator[dict]:
        for dict_template in dict_templates:
            for expression in self.make_expression_list(
                self.sheet.cell(
                    row=row_number, column=dict_template.get("expression")
                ).value
            ):
                expression_dict = {
                    key: (
                        self.sheet.cell(row=row_number, column=value).value
                        if isinstance(value, int)
                        else value
                    )
                    for key, value in dict_template.items()
                    if key != "expression"
                }
                expression_dict["expression"] = expression
                yield expression_dict

    def make_dict(self, dict_templates: list, row_number: int) -> list:
        return [
            {
                key: (
                    self.sheet.cell(row=row_number, column=value).value
                    if isinstance(value, int)
                    else value
                )
                for key, value in dict_template.items()
            }
            for dict_template in dict_templates
        ]


def extract_collection(
    sheet_importer: SheetImporter, sheetinfo: dict, collection: str
) -> Iterator[dict]:
    return (
        sheet_importer.row_to_concept_dict(
            collection, sheet_info=sheetinfo, row_number=row_number
        )
        for row_number in range(2, sheet_importer.sheet.max_row + 1)
    )


@click.command()
@click.option("--lowercase", is_flag=True, help="Lowercase all terms")
@click.option("--comma", is_flag=True, help="Comma separates terms")
@click.argument("filename")
def main(lowercase, comma, filename):
    path = Path(filename)
    workbook = openpyxl.load_workbook(filename)
    template_json = path.with_name(f"{path.stem.lower()}.template.json")
    sheet_infos = json.loads(template_json.read_text())
    collection = sheet_infos.get("collection")
    try:
        data = {
            "collection": asdict(
                COLLECTION_SCHEMA.load(
                    {
                        "name": f"Collection:{collection}",
                        "info": sheet_infos.get("info"),
                        "owner": sheet_infos.get("owner"),
                        "languages": list(
                            {
                                related_expression.get("language")
                                for sheet_info in sheet_infos.get("sheets")
                                for related_expression in sheet_info.get(
                                    "template"
                                ).get("related_expressions")
                            }
                        ),
                    }
                )
            ),
            "concepts": [
                asdict(
                    cleanup_termwiki_page(
                        TERMWIKI_PAGE_SCHEMA.load(
                            {
                                **concept_dict,
                                "title": f"{concept_dict.get('title')}:"
                                f"{collection}_{index}",
                            }
                        )
                    )
                )
                for (index, concept_dict) in enumerate(
                    all_concepts(lowercase, comma, workbook, sheet_infos, collection),
                    start=1,
                )
            ],
        }
    except ValidationError as error:
        message = f"Error in input data\n{error}"
        raise SystemExit(message) from error

    path.with_name(f"{path.stem.lower().replace(' ', '_')}.result.json").write_text(
        json.dumps(
            data,
            indent=2,
            ensure_ascii=False,
        )
    )


def all_concepts(
    lowercase: bool, comma: bool, workbook, sheet_infos: dict, collection: str
) -> Iterator[dict]:
    return (
        concept
        for sheet_info in sheet_infos.get("sheets", {})
        for concept in extract_collection(
            sheet_importer=SheetImporter(
                workbook[sheet_info.get("sheetname")], comma, lowercase
            ),
            sheetinfo=sheet_info.get("template"),
            collection=collection,
        )
    )
