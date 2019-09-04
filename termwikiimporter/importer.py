# -*- coding: utf-8 -*-
#
#   This program is free software: you can redistribute it and/or modify
#   it under the terms of the GNU General Public License as published by
#   the Free Software Foundation, either version 3 of the License, or
#   (at your option) any later version.
#
#   This program is distributed in the hope that it will be useful,
#   but WITHOUT ANY WARRANTY; without even the implied warranty of
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#   GNU General Public License for more details.
#
#   You should have received a copy of the GNU General Public License
#   along with this file. If not, see <http://www.gnu.org/licenses/>.
#
#   Copyright © 2016-2019 The University of Tromsø
#   http://giellatekno.uit.no & http://divvun.no
#
"""Convert term files to termwiki parsable xml."""

import argparse
import os
import re
from pathlib import Path

import yaml
from lxml import etree

import openpyxl
from openpyxl.comments import Comment
from termwikiimporter import check_tw_expressions, lookup, read_termwiki


class ExcelImporter(object):
    """Convert excel files to xml.

    Attributes:
        filename (str): path to the file that should be imported
            to the termwiki
        concepts (list of concepts): all the concepts that have been
            found in filename
    """

    def __init__(self, filename):
        """Initialise the Importer class."""
        self.filename = filename
        self.workbook = openpyxl.load_workbook(self.filename)

    @property
    def resultname(self):
        """Name of the xml output file."""
        return os.path.splitext(self.filename)[0] + '.xml'

    @property
    def fileinfo(self):
        """Parse information about excel files from a yaml file."""
        yamlname = self.filename.replace('.xlsx', '.yaml')
        with open(yamlname) as yamlfile:
            return yaml.load(yamlfile)

    @staticmethod
    def parse_sheet(sheet, pages, info):
        """Parse a sheet in a openpyxl workbook.

        Add concepts to the pages element.
        """
        rowparsers = [
            RowParser(SheetRow(sheet, index), info)
            for index in range(2, sheet.max_row + 1)
        ]
        for rowparser in rowparsers:
            rowparser.parse_row()
            page = etree.SubElement(pages, 'page')
            page.set('title', rowparser.concept.title)
            concept = etree.SubElement(page, 'concept')
            concept.text = str(rowparser.concept)

    def get_concepts(self):
        """Fetch concepts from all sheets in the workbook."""
        pages = etree.Element('pages')
        for sheet_name in self.fileinfo:
            self.parse_sheet(self.workbook[sheet_name], pages,
                             self.fileinfo[sheet_name])

        return pages

    def write_concepts(self):
        """Write concepts to an xml file."""
        pages = self.get_concepts()
        with open(self.resultname, 'w') as to_file:
            to_file.write(
                etree.tostring(pages, pretty_print=True, encoding='unicode'))


class SheetRow(object):
    """Abstract a row in a openpyxl sheet slightly."""
    def __init__(self, sheet, index):
        self.sheet = sheet
        self.index = index

    def __getitem__(self, key):
        return self.sheet.cell(row=self.index, column=key)


class RowParser(object):
    """Parse a row in a openpyxl sheet into a concept."""

    invalid_expression = re.compile(r'[^\w -]+')

    def __init__(self, row, info):
        self.handler = {
            'related_expressions': self.handle_related_expressions,
            'concept_infos': self.handle_concept_infos,
            'source': self.handle_source,
            'main_category': self.handle_maincategory,
            'collection': self.handle_collection
        }
        self.row = row
        self.concept = read_termwiki.Concept()
        self.info = info

    @property
    def related_expressions(self):
        """Return the expressions of this row."""
        return self.concept.related_expressions

    def parse_row(self):
        """Parse a row with the different handlers."""
        for key in self.info:
            self.handler[key]()

    def make_expression_dict(self, lang, expression):
        """Add more details to the expression, if available."""
        expression_dict = {'expression': expression, 'language': lang}

        for key in self.info['related_expressions'][lang]:
            if key not in ['expression']:
                position = int(self.info['related_expressions'][lang][key])
                if position:
                    if self.row[position].value is not None:
                        expression_dict[key] = str(
                            self.row[position].value).strip()
                else:
                    expression_dict[key] = self.info['related_expressions'][
                        lang][key]

        return expression_dict

    def make_error_dict(self, expression, language):
        """Turn the errors found in an expression cell into a comment."""
        expression_comment = {'invalid': False}
        if self.invalid_expression.search(expression):
            expression_comment['invalid'] = True

        expression_comment['possible_dupes'] = {
            hit.replace(' ', '_')
            for hit in lookup.lookup(expression, language)
        }

        lung = check_tw_expressions.LANG2LANG[language]
        expression_comment['possible_typo'] = any([
            lung in check_tw_expressions.ANALYSERS
            and not check_tw_expressions.ANALYSERS[lung].lookup(part)
            for part in expression.split()
        ])

        return expression_comment

    @staticmethod
    def make_error_strings(error_dict):
        """Turn the error dict into a human readable text."""
        error_strings = []
        if error_dict['invalid']:
            # print(': '.join(f'«{x}»: {hex(ord(x))}' for x in expression))
            error_strings.append(
                '\tInvalid expression: '
                'https://satni.uit.no/termwiki/index.php?title='
                'Excel_files_and_termwiki#Invalid_expression'
            )
        else:
            if error_dict['possible_typo']:
                error_strings.append(
                    '\tPossible typo: '
                    'https://satni.uit.no/termwiki/index.php?title='
                    'Excel_files_and_termwiki#Possible_typo'
                )
            if error_dict['possible_dupes']:
                error_strings.append(
                    '\tPossible dupes: '
                    'https://satni.uit.no/termwiki/index.php?title='
                    'Excel_files_and_termwiki#Possiblie_duplicates'
                )
                for hit in error_dict['possible_dupes']:
                    error_strings.append(
                        f'\t\thttps://satni.uit.no/termwiki/index.php?title={hit}'
                    )

        return '\n'.join(error_strings)

    def handle_related_expressions(self):
        """Read expressions from a cell.

        Also gather info about errors found in this cell.
        """
        for lang in self.info['related_expressions']:
            comment_strings = []
            ex_index = self.info['related_expressions'][lang]['expression']
            expressions = self.row[ex_index].value
            if expressions is not None:
                for expression in self.extract_expression(
                        expressions.replace('\n', ' ').replace('\u000a', ' ')):
                    comment_dict = self.make_error_dict(expression, lang)
                    errors = self.make_error_strings(comment_dict)
                    if errors:
                        comment_strings.append(f'{expression}\n{errors}')
                    self.related_expressions.append(
                        self.make_expression_dict(lang, expression))

            if comment_strings:
                comment = Comment('\n'.join(comment_strings), 'Termwiki bot')
                self.row[ex_index].comment = comment

    @staticmethod
    def extract_expression(expression):
        """Turn the content of an expression cell into individual expressions."""
        return [exp.strip() for exp in expression.split(',')]

    def handle_concept_infos(self):
        """Handle concept info found in an Excel cell."""
        for lang in self.info['concept_infos']:
            for key in self.info['concept_infos'][lang]:
                self.concept.data['concept_infos']

    def handle_source(self):
        """Handle a source cell."""
        self.concept.data['source'] = self.row[self.info['source']]

    def handle_maincategory(self):
        """Handle the main category

        Info fetched either from a cell or from the metadata.
        """
        try:
            position = int(self.info['main_category'])
            main_category = str(self.row[position].value).strip()
        except ValueError:
            main_category = self.info['main_category']

        self.concept.title = f'{main_category}:{self.info["collection"]} {self.row.index}'

    def handle_collection(self):
        """Handle info about the collection."""
        if not self.concept.data['concept'].get('collection'):
            self.concept.data['concept']['collection'] = set()
        collection = self.info['collection'] if 'Collection:' in self.info[
            'collection'] else f'Collection:{self.info["collection"]}'
        self.concept.data['concept']['collection'].add(collection)


def parse_options():
    """Parse options given to the script."""
    parser = argparse.ArgumentParser(
        description='Convert files containing terms to TermWiki mediawiki '
        'format')

    parser.add_argument(
        'termfiles',
        nargs='+',
        help='One or more files containing terms. Each file must have a '
        'yaml file that inform how they should be treated.')

    args = parser.parse_args()

    return args


def main():
    """Convert files to termwiki format."""
    args = parse_options()

    for termfile in args.termfiles:
        importer = ExcelImporter(termfile)
        importer.write_concepts()
        excel_path = Path(termfile)
        importer.workbook.save(excel_path.stem + '.report.xlsx')
